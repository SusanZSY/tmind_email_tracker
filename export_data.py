#!/usr/bin/env python3
"""
Export email activity data from Excel to JSON for the dashboard visualization.
Reads from the Email_Events sheet and generates dashboard data.
"""

import json
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_INPUT = "2025 Outreach List.xlsx"
DEFAULT_SHEET = "Email_Events"
DEFAULT_OUTPUT = "data.json"


def normalize(value: Any) -> str:
    return "" if value is None else str(value).strip().lower()


def timestamp_key(value: Any) -> tuple[int, Any]:
    if isinstance(value, datetime):
        return (0, value)
    if value is None:
        return (2, "")
    return (1, str(value))


def event_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value) if value else ""


def event_day(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)).date()
    except ValueError:
        return None


def step_label(value: Any) -> str:
    if value in (None, ""):
        return "No Step"
    if isinstance(value, float) and value.is_integer():
        return f"Step {int(value)}"
    if isinstance(value, int):
        return f"Step {value}"
    text = str(value).strip()
    if text.replace(".", "", 1).isdigit():
        try:
            number = float(text)
            if number.is_integer():
                return f"Step {int(number)}"
        except ValueError:
            pass
    return text if text.lower().startswith("step") else f"Step {text}"


def daterange(start: date, end: date) -> list[date]:
    days = (end - start).days
    return [start + timedelta(days=offset) for offset in range(days + 1)]


def read_activity_rows(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Read email activity rows from Excel file."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' was not found. Available sheets: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(header).strip() for header in header_row if header is not None]

    required_headers = {"Timestamp", "Type", "Email"}
    missing_headers = required_headers - set(headers)
    if missing_headers:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing_headers))}")

    rows: list[dict[str, Any]] = []
    for row_values in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        if not any(value is not None for value in row_values):
            continue
        rows.append(dict(zip(headers, row_values)))

    return headers, rows


def process_dashboard_data(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Process raw rows into dashboard data."""
    activity_types = ["open", "click"]

    daily_counts: dict[str, Counter[str]] = defaultdict(Counter)
    email_daily_counts: dict[str, dict[str, Counter[str]]] = defaultdict(lambda: defaultdict(Counter))
    step_daily_counts: dict[str, dict[str, Counter[str]]] = defaultdict(lambda: defaultdict(Counter))
    type_counts: Counter[str] = Counter()
    email_counts: Counter[str] = Counter()
    email_type_counts: dict[str, Counter[str]] = defaultdict(Counter)
    step_type_counts: dict[str, Counter[str]] = defaultdict(Counter)
    event_days: list[date] = []
    processed_events: list[dict[str, str]] = []

    for row in rows:
        activity_type = normalize(row.get("Type"))
        if activity_type not in {"open", "click"}:
            continue

        email = str(row.get("Email") or "NO_EMAIL").strip()
        step = step_label(row.get("Step"))
        day = event_day(row.get("Timestamp"))
        date_str = day.isoformat() if day else event_date(row.get("Timestamp"))

        if day:
            event_days.append(day)
        if date_str and email != "NO_EMAIL":
            daily_counts[date_str][activity_type] += 1
            email_daily_counts[email][date_str][activity_type] += 1
            step_daily_counts[step][date_str][activity_type] += 1
            processed_events.append(
                {
                    "date": date_str,
                    "type": activity_type,
                    "email": email,
                    "step": step,
                }
            )
        
        type_counts[activity_type] += 1
        email_counts[email] += 1
        email_type_counts[email][activity_type] += 1
        step_type_counts[step][activity_type] += 1

    all_days = daterange(min(event_days), max(event_days)) if event_days else []
    date_range = [day.isoformat() for day in all_days]

    # Build daily data for every calendar day in the event range.
    daily_data = []
    for date_str in date_range:
        day_entry = {"date": date_str, "total": 0}
        for activity_type in activity_types:
            count = daily_counts[date_str].get(activity_type, 0)
            day_entry[activity_type] = count
            day_entry["total"] += count
        daily_data.append(day_entry)

    # Build type data
    type_data = [
        {"type": activity_type.title(), "count": type_counts[activity_type]}
        for activity_type in activity_types
    ]

    # Build top emails data
    top_email_data = [
        {
            "email": email,
            "count": count,
            "open": email_type_counts[email].get("open", 0),
            "click": email_type_counts[email].get("click", 0),
        }
        for email, count in email_counts.most_common(15)
    ]

    # Compact complete grid: every email has one open and one click value for every date.
    email_daily_data = []
    for email, count in email_counts.most_common():
        email_daily_data.append(
            {
                "email": email,
                "total": count,
                "open": email_type_counts[email].get("open", 0),
                "click": email_type_counts[email].get("click", 0),
                "daily": {
                    "open": [
                        email_daily_counts[email][date_str].get("open", 0)
                        for date_str in date_range
                    ],
                    "click": [
                        email_daily_counts[email][date_str].get("click", 0)
                        for date_str in date_range
                    ],
                },
            }
        )

    step_daily_data = []
    for step in sorted(step_type_counts, key=lambda item: (not item.startswith("Step "), item)):
        open_count = step_type_counts[step].get("open", 0)
        click_count = step_type_counts[step].get("click", 0)
        step_daily_data.append(
            {
                "step": step,
                "total": open_count + click_count,
                "open": open_count,
                "click": click_count,
                "daily": {
                    "open": [
                        step_daily_counts[step][date_str].get("open", 0)
                        for date_str in date_range
                    ],
                    "click": [
                        step_daily_counts[step][date_str].get("click", 0)
                        for date_str in date_range
                    ],
                },
            }
        )

    return {
        "daily": daily_data,
        "types": type_data,
        "steps": step_daily_data,
        "topEmails": top_email_data,
        "events": processed_events,
        "dateRange": date_range,
        "emailDaily": email_daily_data,
        "metrics": {
            "totalEvents": len(rows),
            "uniqueEmails": len(email_counts),
            "activeDays": len(date_range),
            "opens": type_counts["open"],
            "clicks": type_counts["click"],
        }
    }


def export_json(input_file: Path, sheet_name: str, output_file: Path) -> None:
    """Read Excel data and export to JSON."""
    print(f"Reading from '{input_file}' / '{sheet_name}'...")
    headers, rows = read_activity_rows(input_file, sheet_name)
    
    print(f"Processing {len(rows)} activity rows...")
    data = process_dashboard_data(rows)
    
    print(f"Writing to '{output_file}'...")
    with open(output_file, 'w') as f:
        json.dump(data, f, indent=2)
    
    print(f"Successfully exported {len(rows)} events to {output_file}")
    print(f"  - Daily entries: {len(data['daily'])}")
    print(f"  - Email/day rows represented: {len(data['emailDaily']) * len(data['dateRange'])}")
    print(f"  - Activity types: {len(data['types'])}")
    print(f"  - Steps: {len(data['steps'])}")
    print(f"  - Top emails: {len(data['topEmails'])}")
    print(f"  - Total events: {data['metrics']['totalEvents']}")


if __name__ == "__main__":
    input_path = Path(DEFAULT_INPUT)
    output_path = Path(DEFAULT_OUTPUT)
    
    if not input_path.exists():
        print(f"Error: '{input_path}' not found")
        exit(1)
    
    export_json(input_path, DEFAULT_SHEET, output_path)
